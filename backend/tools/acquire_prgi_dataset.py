"""Offline merge of manual PRGI CSV/XLSX page exports into one canonical CSV.

Usage: python tools/acquire_prgi_dataset.py merge \
         --input-dir data/raw_prgi_exports --output data/prgi_titles.csv
"""

import argparse
import csv
import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.csv_mapping import (CANONICAL_COLUMNS, get_value, map_headers,
                             read_csv_rows)
from src.normalize import normalize_title


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = ["" if h is None else str(h).strip() for h in next(it)]
    rows = []
    for vals in it:
        row = {}
        for i, h in enumerate(headers):
            row[h] = "" if i >= len(vals) or vals[i] is None else str(vals[i]).strip()
        rows.append(row)
    wb.close()
    return headers, rows


def _iter_source_rows(input_dir):
    files = sorted(
        f for f in os.listdir(input_dir)
        if f.lower().endswith((".csv", ".xlsx"))
    )
    for fname in files:
        path = os.path.join(input_dir, fname)
        try:
            if fname.lower().endswith(".csv"):
                headers, rows = read_csv_rows(path)[:2]
            else:
                headers, rows = _read_xlsx(path)
        except Exception:
            yield fname, None  # malformed source file
            continue
        yield fname, (headers, rows)


def merge(input_dir: str, output_path: str) -> dict:
    seen_reg = set()
    seen_comp = set()
    accepted_rows = []
    source_files = []
    raw_rows = 0
    malformed = 0
    duplicates = 0

    for fname, data in _iter_source_rows(input_dir):
        source_files.append(fname)
        if data is None:
            malformed += 1
            continue
        headers, rows = data
        mapping = map_headers(headers)
        for row in rows:
            raw_rows += 1
            raw_title = get_value(row, mapping, "title")
            if not raw_title:
                malformed += 1
                continue
            norm = normalize_title(raw_title)["normalized"]
            reg = get_value(row, mapping, "registration_number")
            language = get_value(row, mapping, "language")
            state = get_value(row, mapping, "state")
            periodicity = get_value(row, mapping, "periodicity")

            if reg and reg in seen_reg:
                duplicates += 1
                continue
            comp = f"{norm}|{language}|{state}|{periodicity}"
            if comp in seen_comp:
                duplicates += 1
                continue

            if reg:
                seen_reg.add(reg)
            seen_comp.add(comp)

            accepted_rows.append({
                "title": raw_title,
                "registration_number": reg,
                "registration_date": get_value(row, mapping, "registration_date"),
                "language": language,
                "periodicity": periodicity,
                "publisher": get_value(row, mapping, "publisher"),
                "owner": get_value(row, mapping, "owner"),
                "state": state,
                "district": get_value(row, mapping, "district"),
                "status": "",
            })

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CANONICAL_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(accepted_rows)

    return {
        "source_files": source_files,
        "raw_rows": raw_rows,
        "accepted_rows": len(accepted_rows),
        "duplicate_rows": duplicates,
        "malformed_rows": malformed,
        "output_rows": len(accepted_rows),
    }


def main(argv=None):
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    m = sub.add_parser("merge")
    m.add_argument("--input-dir", required=True)
    m.add_argument("--output", required=True)
    args = ap.parse_args(argv)

    if not os.path.isdir(args.input_dir):
        print(f"error: input dir not found: {args.input_dir}", file=sys.stderr)
        return 1

    report = merge(args.input_dir, args.output)

    out = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "data", "reports",
        f"merge_{time.strftime('%Y%m%d_%H%M%S')}.json",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"Source files: {report['source_files']}")
    print(f"Raw rows:     {report['raw_rows']}")
    print(f"Accepted:     {report['accepted_rows']}")
    print(f"Duplicates:   {report['duplicate_rows']}")
    print(f"Malformed:    {report['malformed_rows']}")
    print(f"Output rows:  {report['output_rows']} -> {args.output}")
    print(f"Report:       {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())